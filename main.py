import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.chart import PieChart, Reference

url = 'https://en.wikipedia.org/wiki/The_world%27s_100_most_threatened_species'
tables = pd.read_html(url)

table1 = tables[0]

validtypes = ['Mammal', 'Reptile', 'Amphibian', 'Bird']

table2 = table1[table1['Type'].isin(validtypes)][['Species', 'Common name', 'Type', 'Estimated population']]
print(table2)

type_counts = table2['Type'].value_counts()

mychart = plt.figure(figsize=(8, 6))
plt.pie(type_counts, labels=type_counts.index, autopct='%1.1f%%', startangle=90)
plt.title('Endangered Species by Type')

plt.axis('equal')
plt.savefig('pie_chart.png')

table2.to_excel('EndangeredSpecies.xlsx')

filepath = "C:/Users/chais/PycharmProjects/FinalProject/EndangeredSpecies.xlsx"
imagepath = "C:/Users/chais/PycharmProjects/FinalProject/pie_chart.png"
image = Image(imagepath)
mywrkbook = load_workbook(filepath)
sheet = mywrkbook.active
sheet.add_image(image, 'I1')


mywrkbook.save(filepath)
